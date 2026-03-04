# =============================================================================
# DATA_Updated.xlsx BUILDER (TEMPLATE → MODEL INPUTS) WITH AT MODE SWITCH + PTDF EXPORT
# =============================================================================
# This script builds Data_Updated.xlsx from a template Data.xlsx for the power-system
# optimization pipeline. It:
#
#   - Loads node-level generation and storage capacities from TYNDP-derived
#     Generation_Data.csv (bus_id can be a country node or an Austrian substation).
#
#   - Loads the network topology from Lines_AT_and_Tielines.csv (bus_u / bus_v),
#     rebuilds Exchange_Data (incidence + electrical parameters), and optionally
#     exports a synchronized PTDF matrix aligned exactly to Exchange_Data
#     (lines x nodes, with an explicit slack column = 0).
#
#   - Supports two Austria representations (auto-detected):
#       1) Aggregated AT: a single node "AT"
#       2) Nodal AT: multiple APG substation nodes (from Substations_AT.py and/or
#          nodes observed in inputs), with Control Area mapping back to "AT".
#
#   - Reconstructs technology-specific sheets (Thermal / Hydro / RES / Flexibility)
#     by mapping TYNDP technologies to the model’s internal “Power Plant Type”
#     taxonomy, aggregating capacities by (Type, Node, Control Area), and keeping
#     template parameter columns/structure intact.
#
#   - Preserves template consistency:
#       * Demand_Profiles can be kept from the template and, if AT is nodal,
#         optionally split into substation columns using shares from Substations_AT.py.
#       * Profile sheets (PV, wind, RoR, inflow, storage profiles) remain
#         control-area based and are pruned to the set of considered countries.
#
#   - Adds a robust fallback for template countries that are present in the template
#     universe but missing in the TYNDP capacity extract: those rows are appended
#     back from the template capacity sheets (normalized to single-node countries).
#
# Output:
#   - Data_Updated.xlsx (model-ready input workbook)
#   - Optional: PTDF_Synchronized.csv (aligned to Exchange_Data ordering)
# =============================================================================

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

import importlib.util
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# =============================================================================
# CONFIG
# =============================================================================

@dataclass(frozen=True)
class Config:
    BASE_DIR: Path = Path(r"C:\Users\Lena\Documents\PSS 2030+\Power_System_Models\Model_Paper_1")

    TEMPLATE_XLSX: Path = BASE_DIR / "Data.xlsx"
    OUT_XLSX: Path = BASE_DIR / "Data_Updated.xlsx"

    # Inputs produced by Extracting_and_Mapping_TYNDP_Data.py
    GEN_CSV: Path = BASE_DIR / "Generation_Grid_Data" / "TYNDP_Generation_Data" / "Generation_Data.csv"

    # Combined AT lines + tielines (bus_u/bus_v)
    LINES_CSV: Path = BASE_DIR / "Generation_Grid_Data" / "Grid_Data" / "Lines_AT_and_Tielines.csv"

    # Austrian substation metadata + shares
    SUBSTATIONS_AT_FILE: Path = BASE_DIR / "Substations_AT.py"
    SUBSTATIONS_AT_FALLBACK: Path = Path(r"/mnt/data/Substations_AT.py")

    # Demand handling:
    # True  -> start from template Demand_Profiles and (if AT nodal) split AT column into substations
    USE_TEMPLATE_DEMAND: bool = True

    # PTDF (synchronized export aligned to Exchange_Data)
    CALC_PTDF: bool = True
    SLACK_NODE: str = "SK"          # slack must be a node name present in considered nodes (usually a country)
    PTDF_EXPORT_CSV: Path = BASE_DIR / "PTDF_Synchronized.csv"
    PTDF_PINV_RCOND: float = 1e-10


CFG = Config()


# =============================================================================
# SMALL HELPERS
# =============================================================================

def _norm(x: Any) -> str:
    return " ".join(str(x).replace("\n", " ").replace("\r", " ").split()).strip()

def _norm_lower(x: Any) -> str:
    return _norm(x).lower()

def _is_country_code(s: str) -> bool:
    s2 = str(s).strip().upper()
    return len(s2) == 2 and s2.isalpha()

def _std_country(x: Any) -> str:
    """
    Standardize to 2-letter country code (AT, DE, SI, IT, ...).
    Also handles ENTSO-E style codes like AT00 → AT.
    """
    s = _norm(x).upper()
    if s == "ITSI":
        return "IT"
    if s in {"APG", "AT00"}:
        return "AT"
    if len(s) == 4 and s.endswith("00") and s[:2].isalpha():
        return s[:2]
    if len(s) == 2 and s.isalpha():
        return s
    return s

def _std_node(x: Any) -> str:
    """
    Standardize a *node* identifier.
    - Country-like identifiers are standardized to 2-letter country codes
    - Otherwise preserve as uppercase (for AT substation names)
    """
    s = _norm(x)
    if not s:
        return s
    up = s.upper()
    if _is_country_code(up):
        return _std_country(up)
    if len(up) == 4 and up.endswith("00") and up[:2].isalpha():
        return _std_country(up)
    return up

def _require_cols(df: pd.DataFrame, cols: list[str], label: str) -> None:
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise ValueError(f"{label}: missing columns {missing}. Found: {list(df.columns)}")

def _coerce_numeric(df: pd.DataFrame, cols: list[str]) -> None:
    for c in cols:
        if c not in df.columns:
            df[c] = 0.0
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)

def _read_csv_flexible(path: Path) -> pd.DataFrame:
    for sep in [";", ",", "\t"]:
        try:
            df = pd.read_csv(path, sep=sep, encoding="utf-8")
            if df.shape[1] >= 2:
                return df
        except Exception:
            pass
    return pd.read_csv(path, encoding="utf-8")

def _delete_sheet_if_exists(wb, name: str) -> None:
    if name in wb.sheetnames:
        wb.remove(wb[name])

def _write_df_to_sheet(ws, df: pd.DataFrame) -> None:
    ws.delete_rows(1, ws.max_row)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

def _get_sheet_columns(wb, sheet_name: str) -> list[str]:
    ws = wb[sheet_name]
    header = []
    for cell in ws[1]:
        header.append(_norm(cell.value) if cell.value is not None else "")
    header = [h for h in header if h != ""]
    return header

def _read_sheet_as_df(wb, sheet_name: str) -> pd.DataFrame:
    ws = wb[sheet_name]
    values = list(ws.values)
    if not values:
        return pd.DataFrame()
    header = [_norm(v) if v is not None else "" for v in values[0]]
    data = values[1:]
    df = pd.DataFrame(data, columns=header)
    df = df.loc[:, ~df.columns.duplicated()].copy()
    return df


# =============================================================================
# NEW: Fallback appender for missing TYNDP countries (e.g. LU)
# =============================================================================

def _append_template_fallback_rows(
    *,
    out_df: pd.DataFrame,
    template_df: pd.DataFrame,
    missing_countries: list[str],
    key_cols: list[str],
) -> pd.DataFrame:
    """
    Append rows from template_df for countries in missing_countries, avoiding duplicates.

    Important:
    - Many templates use Node like "LU_1" while Control Area is "LU".
      For fallback, we normalize Node -> Control Area for these countries so that
      the capacity lands on the single node used in the model ("LU").
    """
    if not missing_countries:
        return out_df
    if template_df is None or template_df.empty:
        return out_df
    if out_df is None:
        out_df = pd.DataFrame()

    fb = template_df.copy()

    if "Control Area" not in fb.columns or "Node" not in fb.columns:
        return out_df

    fb["Control Area"] = fb["Control Area"].astype(str).map(_std_country)
    fb["Node"] = fb["Node"].astype(str).map(_std_node)

    fb = fb[fb["Control Area"].isin(missing_countries)].copy()
    if fb.empty:
        return out_df

    # KEY FIX: normalize Node -> Control Area for fallback countries (e.g. LU_1 -> LU)
    fb["Node"] = fb["Control Area"]

    out = pd.concat([out_df, fb], ignore_index=True)

    if all(c in out.columns for c in key_cols):
        out = out.drop_duplicates(subset=key_cols, keep="first").reset_index(drop=True)

    return out


# =============================================================================
# LOAD Substations_AT.py (manual registry)
# =============================================================================

def load_substations_at_objects(cfg: Config) -> tuple[dict, dict]:
    """
    Returns:
      SUBSTATION_LOOKUP, GSK_STATE_TECH_SHARES
    """
    path = cfg.SUBSTATIONS_AT_FILE
    if not path.exists():
        path = cfg.SUBSTATIONS_AT_FALLBACK
    if not path.exists():
        print(f"[WARN] Substations_AT.py not found at: {cfg.SUBSTATIONS_AT_FILE} nor fallback {cfg.SUBSTATIONS_AT_FALLBACK}")
        return {}, {}

    spec = importlib.util.spec_from_file_location("Substations_AT", str(path))
    if spec is None or spec.loader is None:
        print(f"[WARN] Could not load Substations_AT from {path}")
        return {}, {}
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)  # type: ignore

    sub_lookup = getattr(mod, "SUBSTATION_LOOKUP", {})
    gsk = getattr(mod, "GSK_STATE_TECH_SHARES", {})

    if not isinstance(sub_lookup, dict):
        sub_lookup = {}
    if not isinstance(gsk, dict):
        gsk = {}

    return sub_lookup, gsk


# =============================================================================
# AT MODE DETECTION (robust 1-node AT vs nodal AT)
# =============================================================================

def detect_at_mode_and_nodes(
    *,
    gen_nodes: set[str],
    line_nodes: set[str],
    apg_substations: list[str],
) -> tuple[bool, list[str]]:
    """
    Decide whether AT is nodal and return the AT node list.

    Rule:
      - If ANY non-country, non-"AT" node appears in gen or lines => nodal AT.
        Then AT nodes = (APG substations from Substations_AT, if any) ∪ (those seen in inputs)
      - Else => aggregated AT: AT nodes = ["AT"]

    Returns:
      at_is_nodal, at_nodes
    """
    signal = sorted({n for n in (gen_nodes | line_nodes) if (not _is_country_code(n)) and n != "AT"})
    if signal:
        expanded = sorted(set(apg_substations) | set(signal))
        return True, expanded
    return False, ["AT"]


# =============================================================================
# GSK -> substation shares (state shares split equally within state)
# =============================================================================

def build_state_to_substations(sub_lookup: dict, at_substations: list[str]) -> dict[str, list[str]]:
    state_to_subs: dict[str, list[str]] = {}
    for s in at_substations:
        info = sub_lookup.get(s, {})
        st = None
        if isinstance(info, dict):
            st = info.get("state")
        if isinstance(st, str) and st.strip():
            state_to_subs.setdefault(st.strip(), []).append(s)
    return state_to_subs

def substation_shares_from_gsk(
    *,
    group: str,
    at_substations: list[str],
    sub_lookup: dict,
    gsk: dict,
) -> dict[str, float]:
    """
    Convert state shares (gsk[state][group]) into substation shares.
    Default: split a state's share equally across substations in that state.

    If group missing -> fallback uniform.
    Output sums to 1.0 (unless no substations -> {}).
    """
    group = str(group).strip()
    if not at_substations:
        return {}

    state_to_subs = build_state_to_substations(sub_lookup, at_substations)

    raw: dict[str, float] = {s: 0.0 for s in at_substations}

    any_defined = False
    for st, subs in state_to_subs.items():
        share = None
        if isinstance(gsk, dict) and isinstance(gsk.get(st), dict):
            share = gsk[st].get(group)

        if share is None:
            continue

        try:
            st_share = float(share)
        except Exception:
            continue

        any_defined = True
        if st_share <= 0 or not subs:
            continue

        per_sub = st_share / float(len(subs))
        for s in subs:
            raw[s] = raw.get(s, 0.0) + per_sub

    tot = float(sum(raw.values()))
    if (not any_defined) or (tot <= 0):
        n = len(at_substations)
        return {s: (1.0 / n if n else 0.0) for s in at_substations}

    return {s: float(v) / tot for s, v in raw.items()}

def print_gsk_group_sum(gsk: dict, group: str) -> None:
    ssum = 0.0
    n = 0
    for st, d in (gsk or {}).items():
        if isinstance(d, dict) and group in d and d[group] is not None:
            try:
                ssum += float(d[group])
                n += 1
            except Exception:
                pass
    print(f"[GSK] sum over states for '{group}': {ssum:.6f} (states with value: {n})")


# =============================================================================
# Demand_Profiles: split template AT column into substations (only in nodal mode)
# =============================================================================

def split_at_demand_profiles_from_template(
    demand_df: pd.DataFrame,
    *,
    at_substations: list[str],
    at_shares: dict[str, float],
) -> pd.DataFrame:
    """
    Takes a Demand_Profiles dataframe read from template and:
      - finds the AT column (case-insensitive, also AT00)
      - creates one column per AT substation and sets substation_demand = AT_demand * share(sub)
      - drops the original AT column
      - keeps everything else unchanged
    """
    if demand_df is None or demand_df.empty:
        return demand_df

    cols = list(demand_df.columns)

    at_col = None
    for c in cols:
        if _std_country(c) == "AT":
            at_col = c
            break

    if at_col is None:
        print("[WARN] Demand_Profiles: no AT column found in template -> no splitting performed.")
        return demand_df

    if not at_substations:
        print("[WARN] Demand_Profiles: AT nodal mode requested but at_substations is empty -> no splitting performed.")
        return demand_df

    out = demand_df.copy()
    at_vals = pd.to_numeric(out[at_col], errors="coerce").fillna(0.0)

    insert_pos = cols.index(at_col) + 1
    for i, sub in enumerate(at_substations):
        share = float(at_shares.get(sub, 0.0))
        out.insert(insert_pos + i, sub, at_vals * share)

    out = out.drop(columns=[at_col])
    return out


# =============================================================================
# TECH MAPPING
# =============================================================================

def map_thermal_tech_to_type(tech: str) -> tuple[Optional[str], Optional[str]]:
    t = _norm_lower(tech)

    if t in {
        "battery", "dsr",
        "onshore_wind", "offshore_wind",
        "solar_pv", "solar_rooftop", "solar_thermal", "solar_thermal_storage",
        "run_of_river", "reservoir", "pondage",
        "phs_open", "phs_pure",
    }:
        return None, None

    if t in {"other_res_not_defined", "not defined", "splitting not known"}:
        return "OtherRES", "OtherRES"
    if t in {"waste", "small_biomass", "geothermal", "marine"}:
        return "OtherRES", "OtherRES"

    if t.startswith("hydrogen"):
        return "OtherNonRES", "OtherNonRES"
    if t == "nuclear":
        return "Nuclear", "Nuclear"

    if t.startswith("hard coal") or (t.startswith("coal") and "lignite" not in t):
        if "ccs" in t:
            return "Coal3", "Coal"
        if "old" in t:
            return "Coal1", "Coal"
        return "Coal2", "Coal"

    if t.startswith("lignite"):
        if "ccs" in t:
            return "Lignite3", "Lignite"
        if "old" in t:
            return "Lignite1", "Lignite"
        return "Lignite2", "Lignite"

    if t.startswith("gas") or ("natural gas" in t):
        if "ocgt" in t:
            return "Gas3", "Gas"
        if "conventional" in t and "old" in t:
            return "Gas1", "Gas"
        return "Gas2", "Gas"

    if ("heavy oil" in t) or ("light oil" in t) or (t.startswith("oil") and "shale" not in t):
        return "Oil1", "Oil"

    if t in {"other_non_res", "other non res", "other non-res", "othernonres"}:
        return "OtherNonRES", "OtherNonRES"

    return "OtherNonRES", "OtherNonRES"

def map_hydro_tech_to_type(tech: str) -> Optional[str]:
    t = _norm_lower(tech)
    if t in {"pondage", "reservoir"}:
        return "HS"
    if t in {"phs_open", "phs_pure"}:
        return "PHS"
    return None

def map_res_tech_to_type(tech: str) -> Optional[str]:
    t = _norm_lower(tech)
    if t == "run_of_river":
        return "RoR"
    if t in {"solar_pv", "solar_rooftop", "solar_thermal", "solar_thermal_storage"}:
        return "PV"
    if t == "offshore_wind":
        return "WindOff"
    if t == "onshore_wind":
        return "WindOn"
    return None

def map_flex_tech_to_type(tech: str) -> Optional[str]:
    t = _norm_lower(tech)
    if t == "battery":
        return "Battery"
    if t == "dsr":
        return "DSR"
    return None


# =============================================================================
# Template inflow extractors (control-area totals)
# =============================================================================

def extract_phs_inflow_from_template(wb, phs_sheet_name: str) -> pd.DataFrame:
    df0 = _read_sheet_as_df(wb, phs_sheet_name)
    if df0.empty:
        return pd.DataFrame(columns=["Power Plant Type", "Control Area", "Inflow [GWh/a]"])

    inflow_col = next(
        (c for c in df0.columns if ("inflow" in _norm_lower(c)) and ("gwh" in _norm_lower(c))),
        None,
    )
    if inflow_col is None:
        return pd.DataFrame(columns=["Power Plant Type", "Control Area", "Inflow [GWh/a]"])

    needed = ["Power Plant Type", "Control Area"]
    if not all(c in df0.columns for c in needed):
        return pd.DataFrame(columns=["Power Plant Type", "Control Area", "Inflow [GWh/a]"])

    df = df0[["Power Plant Type", "Control Area", inflow_col]].copy()
    df["Control Area"] = df["Control Area"].map(_std_country)
    df["Power Plant Type"] = df["Power Plant Type"].astype(str).map(_norm)

    df[inflow_col] = df[inflow_col].astype(str).str.replace(",", ".", regex=False)
    df[inflow_col] = pd.to_numeric(df[inflow_col], errors="coerce").fillna(0.0)

    out = (
        df.groupby(["Power Plant Type", "Control Area"], as_index=False)[inflow_col]
        .sum()
        .rename(columns={inflow_col: "Inflow [GWh/a]"})
    )
    return out

def extract_res_inflow_from_template(wb, res_sheet_name: str) -> pd.DataFrame:
    df0 = _read_sheet_as_df(wb, res_sheet_name)
    if df0.empty:
        return pd.DataFrame(columns=["Power Plant Type", "Control Area", "Inflow [GWh/a]"])

    inflow_col = next(
        (c for c in df0.columns if "inflow" in _norm_lower(c) and "gwh" in _norm_lower(c)),
        None,
    )
    if inflow_col is None:
        return pd.DataFrame(columns=["Power Plant Type", "Control Area", "Inflow [GWh/a]"])

    needed = {"Power Plant Type", "Control Area", inflow_col}
    if not needed.issubset(df0.columns):
        return pd.DataFrame(columns=["Power Plant Type", "Control Area", "Inflow [GWh/a]"])

    df = df0[["Power Plant Type", "Control Area", inflow_col]].copy()
    df["Power Plant Type"] = df["Power Plant Type"].astype(str).map(_norm)
    df["Control Area"] = df["Control Area"].astype(str).map(_std_country)

    df[inflow_col] = df[inflow_col].astype(str).str.replace(",", ".", regex=False)
    df[inflow_col] = pd.to_numeric(df[inflow_col], errors="coerce").fillna(0.0)

    out = (
        df.groupby(["Power Plant Type", "Control Area"], as_index=False)[inflow_col]
        .sum()
        .rename(columns={inflow_col: "Inflow [GWh/a]"})
    )
    return out


# =============================================================================
# EXCHANGE_DATA BUILDER (COUNTRIES + AT NODES)
# =============================================================================

def build_exchange_data_from_lines(
    lines_df: pd.DataFrame,
    considered_nodes: list[str],
    *,
    bus_to_ca: dict[str, str],
) -> pd.DataFrame:
    """
    Build incidence matrix over considered_nodes using bus_u / bus_v columns.
    Stores 1/x into "Komplexe Leitwerte". Uses s_nom/smax as NTC placeholders.
    """
    base_cols = [
        "Line",
        "NTC A to B [MW]",
        "NTC B to A [MW]",
        "Komplexe Leitwerte",   # 1/x
        "XBorder",
    ]
    node_cols = list(considered_nodes)
    out_cols = base_cols + node_cols

    if lines_df is None or lines_df.empty:
        return pd.DataFrame(columns=out_cols)

    df = lines_df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    bu = next((c for c in df.columns if c.lower() in {"bus_u", "from_bus", "bus0", "bus_from"}), None)
    bv = next((c for c in df.columns if c.lower() in {"bus_v", "to_bus", "bus1", "bus_to"}), None)
    xcol = next((c for c in df.columns if c.lower() in {"x_ohm", "x", "reactance_ohm", "reactance"}), None)
    smax = next((c for c in df.columns if c.lower() in {"smax_mva", "s_nom_mva", "s_nom", "smax", "s_nom_mw"}), None)

    if bu is None or bv is None:
        raise ValueError(
            "Lines file must contain bus_u/bus_v (or compatible names). "
            f"Columns found: {list(df.columns)}"
        )

    df["A"] = df[bu].astype(str).map(_std_node)
    df["B"] = df[bv].astype(str).map(_std_node)

    keep = set(considered_nodes)
    df = df[df["A"].isin(keep) & df["B"].isin(keep) & (df["A"] != df["B"])].copy()
    df = df.reset_index(drop=True)

    # Unique line names
    pair_keys = df.apply(lambda r: "-".join(sorted([r["A"], r["B"]])), axis=1)
    counters: dict[str, int] = {}
    line_names = []
    for pk, a, b in zip(pair_keys.tolist(), df["A"].tolist(), df["B"].tolist()):
        counters[pk] = counters.get(pk, 0) + 1
        k = counters[pk]
        line_names.append(f"{a}{k}-{b}{k}")
    df["Line"] = line_names

    # 1/x
    if xcol is not None:
        x = pd.to_numeric(df[xcol].astype(str).str.replace(",", ".", regex=False), errors="coerce")
        invx = np.where((x.notna()) & (x != 0.0), 1.0 / x, 0.0)
    else:
        invx = np.zeros(len(df), dtype=float)

    # s_nom / smax
    if smax is not None:
        s = pd.to_numeric(df[smax].astype(str).str.replace(",", ".", regex=False), errors="coerce").fillna(0.0).to_numpy(dtype=float)
    else:
        s = np.zeros(len(df), dtype=float)

    out = pd.DataFrame(columns=out_cols, index=range(len(df)))
    out["Line"] = df["Line"].values
    out["NTC A to B [MW]"] = s
    out["NTC B to A [MW]"] = s
    out["Komplexe Leitwerte"] = invx

    # XBorder: 0 for internal lines within the same control area (esp. within AT),
    # 1 for cross-border lines (different control areas).
    A_ca = df["A"].map(lambda n: bus_to_ca.get(n, n if _is_country_code(n) else "AT"))
    B_ca = df["B"].map(lambda n: bus_to_ca.get(n, n if _is_country_code(n) else "AT"))

    out["XBorder"] = (A_ca != B_ca).astype(int)

    for cc in node_cols:
        out[cc] = 0

    # incidence
    for i, r in df.iterrows():
        out.at[i, r["A"]] = -1
        out.at[i, r["B"]] = +1

    return out


# =============================================================================
# PTDF (SYNCHRONIZED TO EXCHANGE_DATA)
# =============================================================================

def build_ptdf_from_exchange_df(
    exchange_df: pd.DataFrame,
    *,
    node_cols: list[str],
    slack_node: str,
    line_col: str = "Line",
    b_col: str = "Komplexe Leitwerte",  # 1/x
    pinv_rcond: float = 1e-10,
) -> pd.DataFrame:
    if exchange_df is None or exchange_df.empty:
        return pd.DataFrame(index=[], columns=node_cols)

    if line_col not in exchange_df.columns:
        raise ValueError(f"PTDF: missing '{line_col}' in Exchange_Data.")
    if b_col not in exchange_df.columns:
        raise ValueError(f"PTDF: missing '{b_col}' in Exchange_Data (expected 1/x).")

    missing_nodes = [c for c in node_cols if c not in exchange_df.columns]
    if missing_nodes:
        raise ValueError(f"PTDF: node columns missing in Exchange_Data: {missing_nodes[:20]}")

    if slack_node not in node_cols:
        raise ValueError(f"PTDF: slack_node '{slack_node}' not found among node_cols.")

    lines = exchange_df[line_col].astype(str).str.strip().tolist()

    # Incidence matrix (L x N)
    A = exchange_df[node_cols].fillna(0.0).to_numpy(dtype=float)

    # Line susceptance b = 1/x
    b = pd.to_numeric(exchange_df[b_col], errors="coerce").fillna(0.0).to_numpy(dtype=float)

    # Build nodal susceptance matrix Bbus = A^T diag(b) A
    Bbus = (A.T * b) @ A  # (N x N)

    n = len(node_cols)
    slack_idx = node_cols.index(slack_node)
    keep_idx = [i for i in range(n) if i != slack_idx]

    # Reduced B matrix (remove slack)
    Bred = Bbus[np.ix_(keep_idx, keep_idx)]

    try:
        Bred_inv = np.linalg.inv(Bred)
    except np.linalg.LinAlgError:
        Bred_inv = np.linalg.pinv(Bred, rcond=pinv_rcond)

    # Reduced incidence
    A_red = A[:, keep_idx]  # (L x (N-1))

    # PTDF = diag(b) A_red Bred^{-1}
    PTDF_red = (A_red @ Bred_inv) * b[:, None]  # (L x (N-1))

    # Reinsert slack column (zeros)
    PTDF_full = np.zeros((len(lines), n), dtype=float)
    PTDF_full[:, keep_idx] = PTDF_red
    PTDF_full[:, slack_idx] = 0.0

    return pd.DataFrame(PTDF_full, index=lines, columns=node_cols)


# =============================================================================
# CLEAN-UP: Prune profile sheets to considered COUNTRIES only
# =============================================================================

def _prune_profile_sheet_to_countries(
    wb,
    sheet_name: str,
    considered_country_nodes: list[str],
    *,
    keep_first_n_cols: int = 1,
) -> None:
    if sheet_name not in wb.sheetnames:
        return

    ws = wb[sheet_name]

    headers = []
    for j, cell in enumerate(ws[1], start=1):
        h = _norm(cell.value) if cell.value is not None else ""
        headers.append((j, h))

    keep_set = set(considered_country_nodes)
    keep_cols = []

    for j, h in headers:
        if j <= keep_first_n_cols:
            keep_cols.append(j)
            continue
        if _std_country(h) in keep_set:
            keep_cols.append(j)

    drop_cols = sorted([j for j, _ in headers if j not in keep_cols], reverse=True)
    for j in drop_cols:
        ws.delete_cols(j)


# =============================================================================
# MAIN BUILD
# =============================================================================

def build_data_updated(cfg: Config = CFG) -> None:
    # -------------------------
    # Input checks
    # -------------------------
    if not cfg.TEMPLATE_XLSX.exists():
        raise FileNotFoundError(f"Template not found: {cfg.TEMPLATE_XLSX}")
    if not cfg.GEN_CSV.exists():
        raise FileNotFoundError(f"Generation CSV not found: {cfg.GEN_CSV}")
    if not cfg.LINES_CSV.exists():
        raise FileNotFoundError(f"Lines CSV not found: {cfg.LINES_CSV}")

    wb = load_workbook(cfg.TEMPLATE_XLSX)

    # -------------------------
    # Load Substations_AT (manual registry)
    # -------------------------
    sub_lookup, gsk = load_substations_at_objects(cfg)
    apg_substations = sorted({_std_node(k) for k in sub_lookup.keys()}) if sub_lookup else []
    print("SUB_LOOKUP size:", len(sub_lookup))
    print("GSK size:", len(gsk))
    print("APG substations (registry):", len(apg_substations))

    # -------------------------
    # Cache template data BEFORE overwriting
    # -------------------------
    if "Demand_Profiles" not in wb.sheetnames:
        raise ValueError("Template missing sheet 'Demand_Profiles'.")
    template_demand_profiles = _read_sheet_as_df(wb, "Demand_Profiles")

    thermal_specific_sheet = "Thermal_Power_Specific_Data"
    if thermal_specific_sheet not in wb.sheetnames:
        raise ValueError(f"Template missing sheet '{thermal_specific_sheet}'")
    if "Thermal_Power_Data" not in wb.sheetnames:
        raise ValueError("Template missing sheet 'Thermal_Power_Data'")

    phs_candidates = ["PHS_Power_Specific_Data", "(P)HS_Power_Specific_Data", "PHS_Specific_Data", "(PHS)_Specific_Data"]
    phs_sheet = next((s for s in phs_candidates if s in wb.sheetnames), None)
    if phs_sheet is None:
        raise ValueError(f"Template missing a PHS/HS sheet. Tried: {phs_candidates}")

    res_candidates = ["RES_Power_Specific_Data", "RES_Specific_Data", "RES_Power_Specific"]
    res_sheet = next((s for s in res_candidates if s in wb.sheetnames), None)
    if res_sheet is None:
        raise ValueError(f"Template missing a RES sheet. Tried: {res_candidates}")

    flex_candidates = ["Flexibility_Specific_Data", "Flexibility_Power_Specific_Data", "Flexibility_Specific"]
    flex_sheet = next((s for s in flex_candidates if s in wb.sheetnames), None)
    if flex_sheet is None:
        raise ValueError(f"Template missing a Flexibility sheet. Tried: {flex_candidates}")

    # Keep template capacity sheets for fallback
    template_thermal_specific_df = _read_sheet_as_df(wb, thermal_specific_sheet)
    template_phs_specific_df = _read_sheet_as_df(wb, phs_sheet)
    template_res_specific_df = _read_sheet_as_df(wb, res_sheet)
    template_flex_specific_df = _read_sheet_as_df(wb, flex_sheet)

    # Inflow totals in template (per Control Area)
    template_phs_inflow_totals = extract_phs_inflow_from_template(wb, phs_sheet)
    template_res_inflow_totals = extract_res_inflow_from_template(wb, res_sheet)

    # -------------------------
    # Load generation (node-level; AT bus_id can be substation OR "AT")
    # -------------------------
    gen = pd.read_csv(cfg.GEN_CSV)
    _require_cols(gen, ["country", "bus_id", "tech"], "Generation_Data")

    gen["country"] = gen["country"].astype(str).map(_std_country)
    gen["bus_id"] = gen["bus_id"].astype(str).map(_std_node)
    gen["tech"] = gen["tech"].astype(str).map(_norm)

    for c in ["p_nom_MW", "p_nom_pump_MW", "e_nom_MWh", "p_nom_storage_MW"]:
        if c not in gen.columns:
            gen[c] = 0.0
    _coerce_numeric(gen, ["p_nom_MW", "p_nom_pump_MW", "e_nom_MWh", "p_nom_storage_MW"])

    gen["p_nom_turb_MW_eff"] = np.maximum(gen["p_nom_MW"].fillna(0.0), gen["p_nom_storage_MW"].fillna(0.0))

    # -------------------------
    # Load lines (bus_u/bus_v)
    # -------------------------
    lines_df = _read_csv_flexible(cfg.LINES_CSV)
    lines_df.columns = [str(c).strip() for c in lines_df.columns]
    bu = next((c for c in lines_df.columns if c.lower() in {"bus_u", "from_bus", "bus0", "bus_from"}), None)
    bv = next((c for c in lines_df.columns if c.lower() in {"bus_v", "to_bus", "bus1", "bus_to"}), None)
    if bu is None or bv is None:
        raise ValueError(f"Lines CSV missing bus_u/bus_v. Found: {list(lines_df.columns)}")

    buses_from_lines = pd.concat(
        [lines_df[bu].astype(str).map(_std_node), lines_df[bv].astype(str).map(_std_node)],
        ignore_index=True
    ).dropna()
    buses_from_lines = [b for b in buses_from_lines.tolist() if b and str(b).upper() != "NAN"]

    # -------------------------
    # Considered country nodes (from template demand columns + gen countries + lines countries)
    # -------------------------
    demand_country_cols: list[str] = []
    for c in template_demand_profiles.columns:
        cc = _std_country(c)
        if _is_country_code(cc):
            demand_country_cols.append(cc)

    gen_country_nodes = sorted(set(gen["country"].astype(str).tolist()))
    line_country_nodes = sorted({n for n in buses_from_lines if _is_country_code(n)})

    considered_country_nodes = sorted(set(demand_country_cols) | set(gen_country_nodes) | set(line_country_nodes))
    if not considered_country_nodes:
        raise ValueError("No country nodes detected.")

    # Ensure AT is present as a country control area for profiles, even if not in template for some reason
    if "AT" not in considered_country_nodes:
        considered_country_nodes.append("AT")
        considered_country_nodes = sorted(set(considered_country_nodes))

    # -------------------------
    # NEW: Detect capacity countries missing in TYNDP gen -> fallback from template
    # -------------------------
    template_universe_countries = sorted(set(demand_country_cols) | set(line_country_nodes) | set(considered_country_nodes))
    gen_countries = sorted(set(gen["country"].astype(str).tolist()))

    missing_caps_countries = sorted(set(template_universe_countries) - set(gen_countries))
    missing_caps_countries = [c for c in missing_caps_countries if c != "AT"]  # AT is handled separately

    if missing_caps_countries:
        print("[WARN] Missing capacity countries in Generation_Data.csv -> fallback from template for:",
              ", ".join(missing_caps_countries))
    else:
        print("[INFO] No missing capacity countries detected (no fallback needed).")

    # -------------------------
    # Detect AT nodal vs aggregated mode
    # -------------------------
    gen_nodes_set = set(gen["bus_id"].astype(str).tolist())
    line_nodes_set = set(buses_from_lines)

    at_is_nodal, at_nodes = detect_at_mode_and_nodes(
        gen_nodes=gen_nodes_set,
        line_nodes=line_nodes_set,
        apg_substations=apg_substations,
    )

    print(f"[INFO] AT mode: {'NODAL (substations)' if at_is_nodal else 'AGGREGATED (single AT node)'}")
    print(f"[INFO] AT nodes used: {len(at_nodes)}")

    # -------------------------
    # Final considered nodes for network:
    # - all non-AT countries as nodes
    # - AT nodes depending on mode
    # -------------------------
    non_at_country_nodes = sorted([cc for cc in considered_country_nodes if cc != "AT"])

    if at_is_nodal:
        considered_node_cols = sorted(non_at_country_nodes + at_nodes)
    else:
        considered_node_cols = sorted(non_at_country_nodes + ["AT"])

    # Slack must exist
    slack = _std_node(cfg.SLACK_NODE)
    if slack not in considered_node_cols:
        raise ValueError(
            f"Slack node '{slack}' not in considered nodes. "
            f"Choose one of: {considered_node_cols[:30]}..."
        )

    # Control area mapping
    bus_to_ca: dict[str, str] = {}
    for n in considered_node_cols:
        bus_to_ca[n] = n if _is_country_code(n) else "AT"

    # -------------------------
    # Demand_Profiles
    # -------------------------
    for s in ["Total_Demand", "Demand_Shares", "RES_Shares"]:
        _delete_sheet_if_exists(wb, s)

    if cfg.USE_TEMPLATE_DEMAND:
        if at_is_nodal:
            print("[INFO] Using template Demand_Profiles and splitting AT to substations.")
            print_gsk_group_sum(gsk, "demand")

            at_demand_shares = substation_shares_from_gsk(
                group="demand",
                at_substations=at_nodes,
                sub_lookup=sub_lookup,
                gsk=gsk,
            )

            demand_out = split_at_demand_profiles_from_template(
                template_demand_profiles,
                at_substations=at_nodes,
                at_shares=at_demand_shares,
            )
            _write_df_to_sheet(wb["Demand_Profiles"], demand_out)
        else:
            print("[INFO] AT is aggregated -> keeping template Demand_Profiles unchanged (no AT splitting).")
            _write_df_to_sheet(wb["Demand_Profiles"], template_demand_profiles)
    else:
        raise NotImplementedError("USE_TEMPLATE_DEMAND=False not implemented in this revision.")

    # -------------------------
    # Thermal_Power_Specific_Data
    # -------------------------
    thermal_params = _read_sheet_as_df(wb, "Thermal_Power_Data")
    _require_cols(thermal_params, ["Power Plant Type"], "Thermal_Power_Data")
    thermal_params = thermal_params.copy()

    mapped = gen["tech"].apply(map_thermal_tech_to_type)
    gen["Power Plant Type"] = mapped.apply(lambda x: x[0])
    gen["Group_from_map"] = mapped.apply(lambda x: x[1])

    thermal = gen.dropna(subset=["Power Plant Type"]).copy()
    thermal["Node"] = thermal["bus_id"]
    thermal["Control Area"] = thermal["bus_id"].map(bus_to_ca).fillna(thermal["country"].map(_std_country))
    thermal["Control Area"] = thermal["Control Area"].astype(str).map(_std_country)
    thermal["Installed Capacity [GW]"] = thermal["p_nom_turb_MW_eff"] / 1000.0

    thermal_agg = thermal.groupby(["Power Plant Type", "Node", "Control Area"], as_index=False)["Installed Capacity [GW]"].sum()
    out_thermal = thermal_agg.merge(thermal_params, on="Power Plant Type", how="left")

    if "Group" in out_thermal.columns:
        grp_map = (
            thermal.drop_duplicates(["Power Plant Type"])[["Power Plant Type", "Group_from_map"]]
            .set_index("Power Plant Type")["Group_from_map"]
            .to_dict()
        )
        out_thermal["Group"] = out_thermal["Group"].fillna(out_thermal["Power Plant Type"].map(grp_map))
    else:
        grp_map = thermal.drop_duplicates(["Power Plant Type"]).set_index("Power Plant Type")["Group_from_map"].to_dict()
        out_thermal["Group"] = out_thermal["Power Plant Type"].map(grp_map)

    thermal_cols = _get_sheet_columns(wb, thermal_specific_sheet)
    missing_cols = [c for c in thermal_cols if c not in out_thermal.columns]
    if missing_cols:
        raise ValueError(
            f"Thermal_Power_Specific_Data: missing columns vs template: {missing_cols}\n"
            f"Available: {list(out_thermal.columns)}"
        )
    out_thermal = out_thermal[thermal_cols].copy()

    # NEW: fallback append from template for countries missing in TYNDP (e.g. LU)
    out_thermal = _append_template_fallback_rows(
        out_df=out_thermal,
        template_df=template_thermal_specific_df[thermal_cols] if not template_thermal_specific_df.empty else template_thermal_specific_df,
        missing_countries=missing_caps_countries,
        key_cols=["Power Plant Type", "Node", "Control Area"],
    )

    _write_df_to_sheet(wb[thermal_specific_sheet], out_thermal)

    # -------------------------
    # (P)HS_Power_Specific_Data (HS/PHS)
    #   - Inflow totals are per control area in template
    #   - If AT nodal: split inflow across substations using gsk groups 'hs_inflow' / 'phs_inflow'
    # -------------------------
    phs_cols = _get_sheet_columns(wb, phs_sheet)

    hydro_type = gen["tech"].apply(map_hydro_tech_to_type)
    hydro = gen.loc[hydro_type.notna()].copy()
    hydro["Power Plant Type"] = hydro["tech"].apply(map_hydro_tech_to_type)

    hydro["Node"] = hydro["bus_id"]
    hydro["Control Area"] = hydro["bus_id"].map(bus_to_ca).fillna(hydro["country"].map(_std_country))
    hydro["Control Area"] = hydro["Control Area"].astype(str).map(_std_country)

    hydro["Turbine Capacity [GW]"] = hydro["p_nom_turb_MW_eff"] / 1000.0
    hydro["Pump Capacity [GW]"] = hydro["p_nom_pump_MW"] / 1000.0
    hydro["Energy (max) [GWh]"] = hydro["e_nom_MWh"] / 1000.0

    hydro_agg = hydro.groupby(["Power Plant Type", "Control Area", "Node"], as_index=False)[
        ["Turbine Capacity [GW]", "Pump Capacity [GW]", "Energy (max) [GWh]"]
    ].sum()

    # merge template inflow totals (per Control Area)
    if not template_phs_inflow_totals.empty:
        infl = template_phs_inflow_totals.copy()
        infl["Control Area"] = infl["Control Area"].map(_std_country)
        infl["Power Plant Type"] = infl["Power Plant Type"].astype(str).map(_norm)
        hydro_agg = hydro_agg.merge(infl, on=["Power Plant Type", "Control Area"], how="left")
    else:
        hydro_agg["Inflow [GWh/a]"] = np.nan

    hydro_agg["Inflow [GWh/a]"] = pd.to_numeric(hydro_agg["Inflow [GWh/a]"], errors="coerce").fillna(0.0)

    if at_is_nodal:
        print_gsk_group_sum(gsk, "hs_inflow")
        print_gsk_group_sum(gsk, "phs_inflow")
        hs_sh = substation_shares_from_gsk(group="hs_inflow", at_substations=at_nodes, sub_lookup=sub_lookup, gsk=gsk)
        phs_sh = substation_shares_from_gsk(group="phs_inflow", at_substations=at_nodes, sub_lookup=sub_lookup, gsk=gsk)

        def _split_inflow_row(r: pd.Series) -> float:
            ca = _std_country(r.get("Control Area"))
            node = _std_node(r.get("Node"))
            ptype = _norm(r.get("Power Plant Type"))
            inflow_total = float(r.get("Inflow [GWh/a]", 0.0))

            if ca != "AT":
                return inflow_total

            if node not in at_nodes:
                return 0.0
            if ptype == "HS":
                return inflow_total * float(hs_sh.get(node, 0.0))
            if ptype == "PHS":
                return inflow_total * float(phs_sh.get(node, 0.0))
            return 0.0

        hydro_agg["Inflow [GWh/a]"] = hydro_agg.apply(_split_inflow_row, axis=1)

    out_hydro = pd.DataFrame(columns=phs_cols)
    for c in phs_cols:
        out_hydro[c] = hydro_agg[c] if c in hydro_agg.columns else np.nan

    # NEW: fallback append from template for countries missing in TYNDP (e.g. LU)
    out_hydro = _append_template_fallback_rows(
        out_df=out_hydro,
        template_df=template_phs_specific_df[phs_cols] if not template_phs_specific_df.empty else template_phs_specific_df,
        missing_countries=missing_caps_countries,
        key_cols=["Power Plant Type", "Node", "Control Area"],
    )

    _write_df_to_sheet(wb[phs_sheet], out_hydro)

    # -------------------------
    # RES_Power_Specific_Data
    #   Profiles stay per control area; inflow parameter stays CA-level concept -> keep at 0 at node level
    # -------------------------
    res_cols = _get_sheet_columns(wb, res_sheet)

    res_type = gen["tech"].apply(map_res_tech_to_type)
    res = gen.loc[res_type.notna()].copy()
    res["Power Plant Type"] = res["tech"].apply(map_res_tech_to_type)

    res["Node"] = res["bus_id"]
    res["Control Area"] = res["bus_id"].map(bus_to_ca).fillna(res["country"].map(_std_country))
    res["Control Area"] = res["Control Area"].astype(str).map(_std_country)
    res["Installed Capacity [GW]"] = res["p_nom_turb_MW_eff"] / 1000.0

    res_agg = res.groupby(["Power Plant Type", "Control Area", "Node"], as_index=False)[["Installed Capacity [GW]"]].sum()

    if "Inflow [GWh/a]" in res_cols:
        res_agg["Inflow [GWh/a]"] = 0.0

    out_res = pd.DataFrame(columns=res_cols)
    for c in res_cols:
        out_res[c] = res_agg[c] if c in res_agg.columns else np.nan

    # NEW: fallback append from template for countries missing in TYNDP (e.g. LU)
    out_res = _append_template_fallback_rows(
        out_df=out_res,
        template_df=template_res_specific_df[res_cols] if not template_res_specific_df.empty else template_res_specific_df,
        missing_countries=missing_caps_countries,
        key_cols=["Power Plant Type", "Node", "Control Area"],
    )

    _write_df_to_sheet(wb[res_sheet], out_res)

    # -------------------------
    # Flexibility_Power_Specific_Data
    # -------------------------
    flex_cols = _get_sheet_columns(wb, flex_sheet)

    flex_type = gen["tech"].apply(map_flex_tech_to_type)
    flex = gen.loc[flex_type.notna()].copy()
    flex["Power Plant Type"] = flex["tech"].apply(map_flex_tech_to_type)

    flex["Node"] = flex["bus_id"]
    flex["Control Area"] = flex["bus_id"].map(bus_to_ca).fillna(flex["country"].map(_std_country))
    flex["Control Area"] = flex["Control Area"].astype(str).map(_std_country)

    flex["Turbine Capacity [GW]"] = flex["p_nom_turb_MW_eff"] / 1000.0
    flex["Pump Capacity [GW]"] = flex["p_nom_pump_MW"] / 1000.0
    flex["Energy (max) [GWh]"] = flex["e_nom_MWh"] / 1000.0

    flex_agg = flex.groupby(["Power Plant Type", "Control Area", "Node"], as_index=False)[
        ["Turbine Capacity [GW]", "Pump Capacity [GW]", "Energy (max) [GWh]"]
    ].sum()

    out_flex = pd.DataFrame(columns=flex_cols)
    for c in flex_cols:
        out_flex[c] = flex_agg[c] if c in flex_agg.columns else np.nan

    # NEW: fallback append from template for countries missing in TYNDP (e.g. LU)
    out_flex = _append_template_fallback_rows(
        out_df=out_flex,
        template_df=template_flex_specific_df[flex_cols] if not template_flex_specific_df.empty else template_flex_specific_df,
        missing_countries=missing_caps_countries,
        key_cols=["Power Plant Type", "Node", "Control Area"],
    )

    _write_df_to_sheet(wb[flex_sheet], out_flex)

    # -------------------------
    # Exchange_Data (REBUILT over countries + AT nodes depending on mode)
    # -------------------------
    exchange_sheet = "Exchange_Data"
    if exchange_sheet not in wb.sheetnames:
        raise ValueError(f"Template missing sheet '{exchange_sheet}'")

    out_exchange = build_exchange_data_from_lines(
        lines_df=lines_df,
        considered_nodes=considered_node_cols,
        bus_to_ca=bus_to_ca,
    )

    _write_df_to_sheet(wb[exchange_sheet], out_exchange)
    print("Exchange_Data rows:", len(out_exchange))
    print("Exchange_Data columns:", list(out_exchange.columns))

    # -------------------------
    # PTDF export (synchronized)
    # -------------------------
    if cfg.CALC_PTDF:
        node_cols = considered_node_cols[:]  # must match Exchange_Data node order

        ptdf_full = build_ptdf_from_exchange_df(
            exchange_df=out_exchange,
            node_cols=node_cols,
            slack_node=slack,
            pinv_rcond=cfg.PTDF_PINV_RCOND,
        )

        cfg.PTDF_EXPORT_CSV.parent.mkdir(parents=True, exist_ok=True)
        ptdf_full.to_csv(cfg.PTDF_EXPORT_CSV, encoding="utf-8")
        print(f"[OK] PTDF exported (synchronized): {cfg.PTDF_EXPORT_CSV}")
        print("PTDF shape (lines x nodes):", ptdf_full.shape, "| slack:", slack)
    else:
        print("[INFO] PTDF calculation/export disabled (CALC_PTDF=False).")

    # -------------------------
    # Clean-up: Restrict inherited profile sheets to considered COUNTRIES only
    # (profiles remain control-area based)
    # -------------------------
    profile_sheets = [
        "PHS_Storage_Profiles",
        "PHS_Inflow_Profiles",
        "HS_Inflow_Profiles",
        "RoR_Profile",
        "WindOn_Profile",
        "WindOff_Profile",
        "PV_Profile",
    ]
    # Always keep AT column for profiles
    countries_for_profiles = sorted(set([_std_country(c) for c in considered_country_nodes] + ["AT"]))

    for sh in profile_sheets:
        _prune_profile_sheet_to_countries(
            wb,
            sh,
            considered_country_nodes=countries_for_profiles,
            keep_first_n_cols=2 if sh in {"PHS_Inflow_Profiles", "HS_Inflow_Profiles"} else 1,
        )

    # -------------------------
    # Save workbook
    # -------------------------
    cfg.OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(cfg.OUT_XLSX)
    print("\nSaved:", cfg.OUT_XLSX)


if __name__ == "__main__":
    build_data_updated()
